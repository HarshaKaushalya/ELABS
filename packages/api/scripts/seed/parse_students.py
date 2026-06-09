"""
Parse 'Lab Schedule - Semester 6 - 24th Batch - 2025_2026.xlsx'
and emit a SQL seed file: 08_semester6_students.sql
"""
import openpyxl, hashlib, re
from pathlib import Path

XLSX = Path(__file__).parent / "Lab Schedule - Semester 6 - 24th Batch - 2025_2026.xlsx"
OUT  = Path(__file__).parent / "08_semester6_students.sql"

wb = openpyxl.load_workbook(str(XLSX), data_only=True)

# ── 1. Read student list ──────────────────────────────────────────────────────
ws_students = wb["Group list - EIE"]
students = []
current_group = None
for row in ws_students.iter_rows(values_only=True):
    reg   = row[1]
    name  = row[2]
    group = row[3]
    if group and str(group).startswith("EE"):
        current_group = str(group).strip()
    if reg and str(reg).startswith("EG/"):
        students.append({
            "reg":   str(reg).strip(),
            "name":  str(name).strip(),
            "group": current_group,
        })

print(f"Students parsed: {len(students)}")

# ── 2. Read schedule ──────────────────────────────────────────────────────────
ws_sched = wb["Lab Schedule"]
rows     = list(ws_sched.iter_rows(values_only=True))

# Row index 2 = module names, row 3 = lab sub-labels
module_row = rows[2]   # 0-indexed
lab_row    = rows[3]

# Build column → (module, lab_label) map
col_map = {}
current_module = None
for i, v in enumerate(module_row):
    if v and i >= 3:
        current_module = str(v).strip()
    lab = lab_row[i] if i < len(lab_row) else None
    if current_module and lab:
        col_map[i] = {"module": current_module, "lab": str(lab).strip()}

# Modules encountered
modules = list({v["module"]: True for v in col_map.values()}.keys())
print(f"Modules: {modules}")

# Schedule entries: list of {date, time_str, module, lab, groups:[]}
schedule_entries = []
for row in rows[4:]:
    date_val = row[1]
    time_str = row[2]
    if not date_val or not time_str or "Holiday" in str(time_str) or "Recess" in str(time_str):
        continue
    try:
        date_str = date_val.strftime("%Y-%m-%d")
    except Exception:
        continue
    for col_i, meta in col_map.items():
        cell_val = row[col_i] if col_i < len(row) else None
        if cell_val:
            # groups like "EE01,EE02" or "RE01,RE02"
            grp_list = [g.strip() for g in str(cell_val).split(",") if g.strip()]
            schedule_entries.append({
                "date":    date_str,
                "time":    str(time_str).strip(),
                "module":  meta["module"],
                "lab":     meta["lab"],
                "groups":  grp_list,
            })

print(f"Schedule entries: {len(schedule_entries)}")

# ── 3. Emit SQL ───────────────────────────────────────────────────────────────
sql_lines = ["USE elabs;\n"]

# Password note: bcrypt of 'elabs2024' inserted via API seed script
# Here we insert a placeholder that the API /admin/students/import will overwrite.
# For direct DB seed we use a known bcrypt hash.
# $2b$10$ ... = bcrypt('elabs2024', 10)
HASH = "$2b$10$YourHashHereReplacedByAPIImport"

# -- semesters
sql_lines.append("-- Semester 6")
sql_lines.append("INSERT IGNORE INTO semesters (id, name, level) VALUES (6, 'Semester 6', 6);\n")

# -- modules (extract code + name from strings like "EE6207 Digital Signal processing")
seen_modules = {}
for m in modules:
    parts = m.split(" ", 1)
    code = parts[0].strip()
    name = parts[1].strip() if len(parts) > 1 else code
    seen_modules[code] = name

sql_lines.append("-- Modules")
for code, name in seen_modules.items():
    safe_name = name.replace("'", "''")
    sql_lines.append(
        f"INSERT IGNORE INTO modules (code, name, semester_id) VALUES ('{code}', '{safe_name}', 6);"
    )
sql_lines.append("")

# -- student_profiles table DDL (idempotent)
sql_lines.append("""-- Student profiles table
CREATE TABLE IF NOT EXISTS student_profiles (
  user_id      BIGINT NOT NULL PRIMARY KEY,
  reg_number   VARCHAR(30) NOT NULL UNIQUE,
  group_code   VARCHAR(20),
  semester     INT DEFAULT 6,
  department   VARCHAR(80) DEFAULT 'Electrical and Information Engineering',
  must_change_password TINYINT(1) DEFAULT 1,
  FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- Add must_change_password column to users if missing
ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password TINYINT(1) DEFAULT 0;
""")

# -- timetable_slots table DDL (idempotent)
sql_lines.append("""-- Timetable slots table
CREATE TABLE IF NOT EXISTS timetable_slots (
  id           BIGINT AUTO_INCREMENT PRIMARY KEY,
  session_date DATE NOT NULL,
  time_slot    VARCHAR(30) NOT NULL,
  module_code  VARCHAR(20) NOT NULL,
  lab_label    VARCHAR(20) NOT NULL,
  group_code   VARCHAR(20) NOT NULL,
  academic_year VARCHAR(20) DEFAULT '2025/2026',
  INDEX idx_timetable_group (group_code),
  INDEX idx_timetable_date  (session_date)
);
""")

# -- messages & notifications tables DDL
sql_lines.append("""-- Messages table
CREATE TABLE IF NOT EXISTS broadcast_messages (
  id           BIGINT AUTO_INCREMENT PRIMARY KEY,
  sender_id    BIGINT NOT NULL,
  subject      VARCHAR(255),
  body         TEXT NOT NULL,
  target_type  ENUM('ALL','GROUP','USER') NOT NULL DEFAULT 'ALL',
  target_group VARCHAR(20) NULL,
  target_user  BIGINT NULL,
  created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
  FOREIGN KEY (sender_id) REFERENCES users(id) ON DELETE CASCADE,
  FOREIGN KEY (target_user) REFERENCES users(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS message_recipients (
  message_id   BIGINT NOT NULL,
  user_id      BIGINT NOT NULL,
  is_read      TINYINT(1) DEFAULT 0,
  read_at      DATETIME NULL,
  PRIMARY KEY (message_id, user_id),
  FOREIGN KEY (message_id) REFERENCES broadcast_messages(id) ON DELETE CASCADE,
  FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
);

-- Notifications table
CREATE TABLE IF NOT EXISTS notifications (
  id           BIGINT AUTO_INCREMENT PRIMARY KEY,
  user_id      BIGINT NOT NULL,
  type         ENUM('BORROW_APPROVED','BORROW_OVERDUE','BORROW_RETURNED','LAB_REMINDER','FIRE_ALERT','BROADCAST','SYSTEM') NOT NULL,
  title        VARCHAR(255) NOT NULL,
  body         TEXT,
  is_read      TINYINT(1) DEFAULT 0,
  read_at      DATETIME NULL,
  meta         JSON NULL,
  created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
  FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
  INDEX idx_notifications_user (user_id),
  INDEX idx_notifications_unread (user_id, is_read)
);
""")

# -- Insert students
sql_lines.append("-- 75 Semester 6 EIE students (password = 'elabs2024', must change on first login)")
sql_lines.append("-- NOTE: password_hash below is a placeholder. Run `node scripts/seed/hash_and_insert.js` to insert with real bcrypt hash.")
sql_lines.append("-- Or use the Admin → Import Students UI which hashes properly.")
sql_lines.append("")

for s in students:
    reg_no = s["reg"]
    name   = s["name"].replace("'", "''")
    group  = s["group"] or "EE01"
    # email: eg2022-4904@eng.ruh.ac.lk derived from reg no
    reg_num = reg_no.replace("EG/2022/", "").replace("/", "-")
    email   = f"eg{reg_num.lower()}@eng.ruh.ac.lk"
    sql_lines.append(f"-- {reg_no} | {s['name']} | {group}")
    sql_lines.append(
        f"SET @uid_{reg_num.replace('-','_')} = NULL;\n"
        f"INSERT IGNORE INTO users (index_no, full_name, email, password_hash, must_change_password, is_active)\n"
        f"  VALUES ('{reg_no}', '{name}', '{email}', 'PLACEHOLDER_HASH_elabs2024', 1, 1);\n"
        f"SET @uid_{reg_num.replace('-','_')} = LAST_INSERT_ID();\n"
        f"INSERT IGNORE INTO user_roles (user_id, role_id)\n"
        f"  SELECT @uid_{reg_num.replace('-','_')}, id FROM roles WHERE name='STUDENT' LIMIT 1;\n"
        f"INSERT IGNORE INTO student_profiles (user_id, reg_number, group_code, semester, must_change_password)\n"
        f"  SELECT @uid_{reg_num.replace('-','_')}, '{reg_no}', '{group}', 6, 1\n"
        f"  WHERE @uid_{reg_num.replace('-','_')} IS NOT NULL AND @uid_{reg_num.replace('-','_')} > 0;\n"
    )

sql_lines.append("")

# -- Timetable slots
sql_lines.append("-- Timetable schedule entries")
sql_lines.append("DELETE FROM timetable_slots WHERE academic_year='2025/2026';")
for e in schedule_entries:
    for grp in e["groups"]:
        safe_module = e["module"].replace("'", "''")
        sql_lines.append(
            f"INSERT INTO timetable_slots (session_date, time_slot, module_code, lab_label, group_code, academic_year)\n"
            f"  VALUES ('{e['date']}', '{e['time']}', '{e['module'].split()[0]}', '{e['lab']}', '{grp}', '2025/2026');"
        )

sql_lines.append("")
sql_lines.append("-- End of Semester 6 seed")

OUT.write_text("\n".join(sql_lines), encoding="utf-8")
print(f"SQL written to: {OUT}")
print(f"Students: {len(students)}, Schedule rows: {len(schedule_entries)}, Modules: {len(seen_modules)}")
